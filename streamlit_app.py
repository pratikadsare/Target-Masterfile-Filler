
import io
import re
from io import BytesIO
import difflib
import pandas as pd
import streamlit as st
import streamlit.components.v1 as components
from typing import List, Dict, Optional
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

st.set_page_config(page_title="Target Masterfile Filler — Interactive Mapping", layout="wide")

# ---- Keep scroll position across reruns ----
def inject_scroll_restoration():
    components.html(
        """
        <script>
        const KEY = 'scrollY';
        window.addEventListener('load', function() {
          const y = sessionStorage.getItem(KEY);
          if (y !== null) { window.scrollTo(0, parseFloat(y)); }
        });
        window.addEventListener('beforeunload', function() {
          sessionStorage.setItem(KEY, window.scrollY);
        });
        </script>
        """,
        height=0
    )

inject_scroll_restoration()

# --- Header ---
st.markdown("<h1 style='text-align: center;'>🧩 Target Masterfile Filler — Interactive Mapping</h1>", unsafe_allow_html=True)
st.markdown("<h4 style='text-align: center; font-style: italic;'>Innovating with AI Today ⏐ Leading Automation Tomorrow</h4>", unsafe_allow_html=True)
st.caption("First sheet only, row 1 = headers, row 2 preserved, data written from row 3. Other sheets unchanged. Duplicates in Partner SKU/Barcode highlighted.")

# ----------------- Helpers -----------------
def _norm_key(s: str) -> str:
    s = str(s or "").strip().lower()
    s = s.replace("&", "and")
    return re.sub(r"[^a-z0-9]+", "", s)

def get_template_headers_from_first_sheet(uploaded_template) -> List[str]:
    try:
        wb = load_workbook(filename=BytesIO(uploaded_template.getbuffer()), data_only=False, keep_vba=uploaded_template.name.lower().endswith(".xlsm"))
        ws = wb.worksheets[0]
        headers = []
        max_col = ws.max_column or 1
        for c in range(1, max_col + 1):
            v = ws.cell(row=1, column=c).value
            if v is None:
                continue
            s = str(v).strip()
            if s:
                headers.append(s)
        return headers
    except Exception as e:
        st.error(f"Failed to read template headers: {e}")
        return []

def read_raw(uploaded_file, sheet: Optional[str]) -> pd.DataFrame:
    if uploaded_file is None:
        return pd.DataFrame()
    name = uploaded_file.name.lower()
    try:
        if name.endswith((".csv", ".txt")):
            return pd.read_csv(uploaded_file)
        if sheet is not None:
            return pd.read_excel(uploaded_file, sheet_name=sheet, engine="openpyxl")
        xls = pd.ExcelFile(uploaded_file, engine="openpyxl")
        first = xls.sheet_names[0]
        return pd.read_excel(xls, sheet_name=first, engine="openpyxl")
    except Exception as e:
        st.error(f"Failed to read RAW: {e}")
        return pd.DataFrame()

def build_header_index_first_sheet(ws, header_row: int = 1) -> Dict[str, int]:
    headers = {}
    max_col = ws.max_column or 1
    for c in range(1, max_col + 1):
        v = ws.cell(row=header_row, column=c).value
        if v is None:
            continue
        key = _norm_key(v)
        if key and key not in headers:
            headers[key] = c
    return headers

def highlight_duplicates(ws, header_map: Dict[str, int], header_labels: List[str], start_row: int = 3):
    dup_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    for hdr in header_labels:
        key = _norm_key(hdr)
        col_idx = header_map.get(key)
        if not col_idx:
            continue
        counts = {}
        max_row = ws.max_row or start_row
        for r in range(start_row, max_row + 1):
            v = ws.cell(row=r, column=col_idx).value
            if v is None: 
                continue
            s = str(v).strip()
            if not s:
                continue
            counts[s.upper()] = counts.get(s.upper(), 0) + 1
        if counts:
            for r in range(start_row, max_row + 1):
                v = ws.cell(row=r, column=col_idx).value
                if v is None:
                    continue
                s = str(v).strip()
                if not s:
                    continue
                if counts.get(s.upper(), 0) > 1:
                    ws.cell(row=r, column=col_idx).fill = dup_fill

def fill_first_sheet_by_headers(template_bytes: BytesIO, mapping_df: pd.DataFrame, raw_df: pd.DataFrame, template_filename: str) -> BytesIO:
    keep_vba = template_filename.lower().endswith(".xlsm")
    wb = load_workbook(filename=template_bytes, data_only=False, keep_vba=keep_vba)
    ws = wb.worksheets[0]  # FIRST sheet only

    tpl_header_to_col = build_header_index_first_sheet(ws, header_row=1)
    if not tpl_header_to_col:
        raise ValueError("No headers found in row 1 of the first sheet. Please ensure row 1 contains headers.")

    raw_norm = {c.strip().lower(): c for c in raw_df.columns}

    pairs = []
    missing_raw, missing_tpl = [], []
    for _, r in mapping_df.iterrows():
        raw_hdr_lc = str(r["raw_header"]).strip().lower()
        tpl_hdr_norm = _norm_key(r["template_header"])
        if not raw_hdr_lc:
            continue
        raw_col_name = raw_norm.get(raw_hdr_lc)
        col_idx = tpl_header_to_col.get(tpl_hdr_norm)
        if raw_col_name is None:
            missing_raw.append(r["raw_header"]); continue
        if col_idx is None:
            missing_tpl.append(r["template_header"]); continue
        pairs.append((raw_col_name, col_idx))

    if missing_tpl:
        st.warning("Template headers not found in row 1 (skipped): " + ", ".join(sorted(set(missing_tpl))))
    if missing_raw:
        st.warning("RAW columns missing (skipped): " + ", ".join(sorted(set(missing_raw))))

    start_row = 3
    for i, (_, raw_row) in enumerate(raw_df.iterrows()):
        out_row = start_row + i
        for raw_col_name, col_idx in pairs:
            val = raw_row[raw_col_name]
            ws.cell(row=out_row, column=col_idx, value=("" if pd.isna(val) else val))

    highlight_duplicates(ws, tpl_header_to_col, header_labels=["Partner SKU", "Barcode"], start_row=start_row)

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out

def sanitize_filename(name: str, keep_xlsm: bool) -> str:
    name = (name or "").strip()
    if not name:
        return "filled_masterfile.xlsm" if keep_xlsm else "filled_masterfile.xlsx"
    name = re.sub(r'[\\/*?:"<>|]+', "_", name)
    if keep_xlsm and not name.lower().endswith(".xlsm"):
        name += ".xlsm"
    elif not keep_xlsm and not name.lower().endswith(".xlsx"):
        name += ".xlsx"
    return name

# ----------------- Session keys -----------------
def _init_state():
    for k, v in {
        "template_file": None,
        "template_headers": [],
        "template_sig": "",
        "raw_df": pd.DataFrame(),
        "raw_headers": [],
        "raw_sig": "",
        "mapping_table": pd.DataFrame(columns=["Template","Raw"]),
        "uniqueness_mode": "Validate on Download",
        "auto_resolve_on_download": False,
        "edit_tick": 0,
        "row_edit_tick": {},  # row index -> last edit tick
    }.items():
        if k not in st.session_state:
            st.session_state[k] = v

_init_state()

# ----------------- UI Tabs -----------------
tab1, tab2, tab3 = st.tabs([
    "1) Upload Masterfile Template",
    "2) Upload Raw Data",
    "3) Interactive Mapping & Download"
])

# ---- Tab 1: Template ----
with tab1:
    st.subheader("Upload Masterfile Template (XLSX or XLSM)")
    template_file = st.file_uploader("Masterfile (Excel .xlsx or .xlsm)", type=["xlsx","xlsm"], key="template_file_upl")

    if template_file is not None:
        template_headers = get_template_headers_from_first_sheet(template_file)
        new_sig = "|".join(template_headers)
        st.success(f"Template loaded. First sheet headers: {len(template_headers)} found.")
        st.dataframe(pd.DataFrame({"Template Headers (row 1)": template_headers}), use_container_width=True)

        # Only rebuild mapping grid if template headers actually changed
        if new_sig != st.session_state.template_sig:
            prev_map = {}
            if not st.session_state.mapping_table.empty:
                for _, r in st.session_state.mapping_table.iterrows():
                    prev_map[str(r["Template"])] = str(r["Raw"]) if pd.notna(r["Raw"]) else ""
            rows = [{"Template": th, "Raw": prev_map.get(th, "")} for th in template_headers]
            st.session_state.mapping_table = pd.DataFrame(rows, columns=["Template","Raw"])
            st.session_state.template_headers = template_headers
            st.session_state.template_sig = new_sig

        st.session_state.template_file = template_file

# ---- Tab 2: Raw ----
with tab2:
    st.subheader("Upload Raw Data (CSV/XLSX)")
    raw_file = st.file_uploader("Raw file", type=["csv","xlsx"], key="raw_file_upl")
    raw_sheet = None
    if raw_file is not None and raw_file.name.lower().endswith(".xlsx"):
        try:
            sheets = pd.ExcelFile(raw_file, engine="openpyxl").sheet_names
            if sheets:
                raw_sheet = st.selectbox("Select RAW sheet", options=sheets, index=0, key="raw_sheet_select")
        except Exception as e:
            st.error(f"Could not read RAW Excel: {e}")

    raw_df = read_raw(raw_file, raw_sheet)
    if not raw_df.empty:
        new_raw_sig = "|".join(list(raw_df.columns.astype(str)))
        st.session_state.raw_df = raw_df
        st.session_state.raw_headers = list(raw_df.columns.astype(str))

        # Clean mapped Raw values that no longer exist when raw headers change
        if new_raw_sig != st.session_state.raw_sig and not st.session_state.mapping_table.empty:
            mt = st.session_state.mapping_table.copy()
            mt["Raw"] = mt["Raw"].apply(lambda x: x if str(x) in st.session_state.raw_headers else "")
            st.session_state.mapping_table = mt
            st.session_state.raw_sig = new_raw_sig

        st.success(f"RAW loaded. Columns: {len(st.session_state.raw_headers)} found.")
        st.dataframe(pd.DataFrame({"Raw Headers": st.session_state.raw_headers}), use_container_width=True)

# ---- Tab 3: Interactive Mapping & Download ----
with tab3:
    st.subheader("Interactive Mapping")
    if not st.session_state.template_headers:
        st.info("Upload a Template in Tab 1 to start.")
    if st.session_state.raw_df.empty:
        st.info("Upload RAW data in Tab 2 to proceed.")

    # Uniqueness mode controls
    st.markdown("**Uniqueness mode**")
    st.session_state.uniqueness_mode = st.radio(
        label="How to handle duplicate Raw selections while editing?",
        options=["Validate on Download", "Live (latest edit wins)"],
        index=0 if st.session_state.uniqueness_mode == "Validate on Download" else 1,
        horizontal=True,
        help=(
            "Validate on Download: no auto clearing while editing; duplicates are checked when you click Process.\n"
            "Live: enforce 1:1 immediately; your latest edit keeps the value and other rows using it are cleared."
        )
    )
    if st.session_state.uniqueness_mode == "Validate on Download":
        st.session_state.auto_resolve_on_download = st.checkbox(
            "Auto‑resolve duplicates on download (keep your latest edits, clear others)",
            value=st.session_state.auto_resolve_on_download
        )

    col_left, col_mid, col_right = st.columns([1, 2, 1], gap="large")

    # Compute status for current mapping
    def compute_status(df: pd.DataFrame) -> pd.DataFrame:
        df = df.copy()
        raw_counts = {}
        for _, r in df.iterrows():
            rv = str(r["Raw"]).strip()
            if rv:
                raw_counts[rv] = raw_counts.get(rv, 0) + 1
        statuses = []
        for _, r in df.iterrows():
            rv = str(r["Raw"]).strip()
            if not rv:
                statuses.append("🔴 Unmapped")
            elif raw_counts.get(rv, 0) > 1:
                statuses.append("🟡 Duplicate")
            else:
                statuses.append("🟢 Mapped")
        df["Status"] = statuses
        return df

    mapped_df = compute_status(st.session_state.mapping_table)

    with col_left:
        st.markdown("**Template Headers (row 1)**")
        # Render a custom HTML list with colors (green for mapped, amber for duplicate, light red for unmapped)
        styles = """
        <style>
        .tpl-item { padding:6px 10px; margin:2px 0; border-radius:6px; font-size:0.92rem; }
        .tpl-mapped { background:#E6FFED; }      /* green tint */
        .tpl-dup { background:#FFF7CC; }         /* soft amber */
        .tpl-unmapped { background:#FFECEC; }    /* soft red */
        .tpl-wrap { max-height:460px; overflow:auto; border:1px solid #eee; border-radius:8px; padding:6px; }
        </style>
        """
        html = [styles, "<div class='tpl-wrap'>"]
        for _, row in mapped_df.iterrows():
            status = row["Status"]
            cls = "tpl-unmapped"
            if status.startswith("🟢"):
                cls = "tpl-mapped"
            elif status.startswith("🟡"):
                cls = "tpl-dup"
            html.append(f"<div class='tpl-item {cls}'>{row['Template']}</div>")
        html.append("</div>")
        st.markdown("".join(html), unsafe_allow_html=True)

    with col_right:
        st.markdown("**Raw Headers**")
        st.dataframe(pd.DataFrame({"Raw": st.session_state.raw_headers}), use_container_width=True, height=420)

        st.markdown("---")
        st.markdown("**Tools**")
        # Auto-map exact
        if st.button("Auto‑map (exact names)"):
            mt = st.session_state.mapping_table.copy()
            raw_headers = st.session_state.raw_headers
            raw_norm_map = {_norm_key(h): h for h in raw_headers}
            for i, row in mt.iterrows():
                if str(row["Raw"]):
                    continue
                tpl = str(row["Template"])
                candidate = raw_norm_map.get(_norm_key(tpl))
                if candidate:
                    mt.at[i, "Raw"] = candidate
            st.session_state.mapping_table = mt
            mapped_df = compute_status(st.session_state.mapping_table)
            st.toast("Exact auto‑map applied.", icon="✅")

        # Auto-map fuzzy
        fuzz_thresh = st.slider("Fuzzy match threshold", min_value=0, max_value=100, value=80, step=1, help="Percent similarity; higher = stricter")
        if st.button("Auto‑map (fuzzy)"):
            mt = st.session_state.mapping_table.copy()
            raw_headers = st.session_state.raw_headers
            for i, row in mt.iterrows():
                if str(row["Raw"]):
                    continue
                tpl = str(row["Template"])
                tpl_norm = _norm_key(tpl)
                candidates = [(h, difflib.SequenceMatcher(None, tpl_norm, _norm_key(h)).ratio()) for h in raw_headers]
                candidates.sort(key=lambda x: x[1], reverse=True)
                if candidates and int(candidates[0][1] * 100) >= fuzz_thresh:
                    best = candidates[0][0]
                    mt.at[i, "Raw"] = best
            st.session_state.mapping_table = mt
            mapped_df = compute_status(st.session_state.mapping_table)
            st.toast("Fuzzy auto‑map applied.", icon="✅")

        # Import mapping
        st.markdown("---")
        imp = st.file_uploader("Import mapping (CSV/XLSX)", type=["csv","xlsx"], key="import_map_upl")
        if imp is not None:
            try:
                if imp.name.lower().endswith(".csv"):
                    imp_df = pd.read_csv(imp)
                else:
                    imp_df = pd.read_excel(imp, engine="openpyxl")
                cols_norm = {_norm_key(c): c for c in imp_df.columns}
                tpl_col = next((cols_norm[c] for c in ["template","templateheader","headerofmasterfiletemplate","target","to"] if c in cols_norm), None)
                raw_col = next((cols_norm[c] for c in ["raw","rawheader","headerofrowsheet","source","from"] if c in cols_norm), None)
                if not tpl_col or not raw_col:
                    st.warning("Could not find 'Template' and 'Raw' columns in the import file.")
                else:
                    mt = pd.DataFrame({"Template": st.session_state.template_headers})
                    imp_small = imp_df[[tpl_col, raw_col]].dropna()
                    imp_small.columns = ["Template","Raw"]
                    imp_small["Raw"] = imp_small["Raw"].astype(str).apply(lambda x: x if x in st.session_state.raw_headers else "")
                    merged = pd.merge(mt, imp_small, on="Template", how="left")
                    merged["Raw"] = merged["Raw"].fillna("")
                    st.session_state.mapping_table = merged
                    mapped_df = compute_status(st.session_state.mapping_table)
                    st.success("Mapping imported into the grid.")
            except Exception as e:
                st.error(f"Failed to import mapping: {e}")

        # Export mapping
        if st.button("Export mapping (CSV)"):
            mt = st.session_state.mapping_table.copy()
            out = mt.to_csv(index=False).encode("utf-8")
            st.download_button("Download mapping.csv", data=out, file_name="mapping.csv", mime="text/csv")

    with col_mid:
        st.markdown("**Center Mapping: pick a Raw column for each Template row**")

        if st.session_state.mapping_table.empty:
            st.info("Mapping grid will appear once both Template and Raw are uploaded.")
        else:
            # Build display df with Status column
            display_df = compute_status(st.session_state.mapping_table)

            # Snapshot BEFORE edit (for change detection)
            prev_df = st.session_state.mapping_table.copy()

            options = [""] + st.session_state.raw_headers
            cfg = {
                "Template": st.column_config.Column(disabled=True),
                "Raw": st.column_config.SelectboxColumn("Raw", options=options, help="Select the raw column that maps to this template header."),
                "Status": st.column_config.Column(disabled=True),
            }

            edited_display = st.data_editor(
                display_df,
                column_config=cfg,
                hide_index=True,
                use_container_width=True,
                num_rows="fixed",
                key="mapping_editor"
            ).copy()

            # Extract the two columns we actually store
            edited = edited_display[["Template","Raw"]].copy()
            edited["Raw"] = edited["Raw"].fillna("").astype(str)

            # Track latest edited rows for "latest wins" logic later
            changed_idxs = [i for i in edited.index if str(edited.loc[i, "Raw"]) != str(prev_df.loc[i, "Raw"])]
            if changed_idxs:
                st.session_state.edit_tick += 1
                for i in changed_idxs:
                    st.session_state.row_edit_tick[i] = st.session_state.edit_tick

            # Live uniqueness enforcement only if selected
            if st.session_state.uniqueness_mode == "Live (latest edit wins)":
                # Keep the last-edited row; clear others that use the same raw
                for i in changed_idxs:
                    v = edited.loc[i, "Raw"].strip()
                    if not v:
                        continue
                    dup_idxs = [j for j in edited.index if j != i and edited.loc[j, "Raw"].strip() == v]
                    for j in dup_idxs:
                        edited.loc[j, "Raw"] = ""

            # Save back
            st.session_state.mapping_table = edited

        st.markdown("---")
        default_name = "filled_masterfile.xlsm" if (st.session_state.template_file and st.session_state.template_file.name.lower().endswith(".xlsm")) else "filled_masterfile.xlsx"
        output_name = st.text_input("Output file name", value=default_name, help="Enter the name to use for the downloaded file (extension will be enforced).")

        can_process = (
            st.session_state.template_file is not None and
            not st.session_state.raw_df.empty and
            not st.session_state.mapping_table.empty
        )

        if not can_process:
            st.info("Please upload Template (Tab 1), Raw (Tab 2), and complete the mapping (Tab 3).")

        if st.button("⚙️ Process & Download", type="primary", disabled=not can_process):
            try:
                mt = st.session_state.mapping_table.copy()
                mt = mt[(mt["Raw"].astype(str).str.strip() != "") & (mt["Template"].astype(str).str.strip() != "")]

                # Duplicate validation if Validate-on-Download
                if st.session_state.uniqueness_mode == "Validate on Download":
                    counts = mt["Raw"].value_counts()
                    dup_raws = [raw for raw, cnt in counts.items() if raw and cnt > 1]
                    if dup_raws:
                        if st.session_state.auto_resolve_on_download:
                            # Resolve duplicates using last-edited wins (fall back to lower row index)
                            keep_set = set()
                            for raw in dup_raws:
                                rows = mt.index[mt["Raw"] == raw].tolist()
                                # Choose row with max edit_tick, else max index
                                best = max(rows, key=lambda r: st.session_state.row_edit_tick.get(r, -1))
                                keep_set.add(best)
                                for r in rows:
                                    if r != best:
                                        mt.at[r, "Raw"] = ""  # clear duplicates
                            # Re-drop cleared rows
                            mt = mt[(mt["Raw"].astype(str).str.strip() != "")]
                            st.toast("Duplicates auto‑resolved (latest edits kept).", icon="✅")
                        else:
                            details = []
                            for raw in dup_raws:
                                rows = mt.index[mt["Raw"] == raw].tolist()
                                names = [f"{mt.at[r, 'Template']}" for r in rows]
                                details.append(f"- **{raw}** → " + ", ".join(names))
                            st.error("Duplicate Raw selections detected. Resolve them or enable 'Auto‑resolve duplicates on download'.")
                            st.markdown("\n".join(details))
                            st.stop()

                writer_map = mt[["Raw","Template"]].copy()
                writer_map.columns = ["raw_header","template_header"]

                if writer_map.empty:
                    st.error("No valid mapping rows selected.")
                else:
                    out_bytes = fill_first_sheet_by_headers(
                        template_bytes=BytesIO(st.session_state.template_file.getbuffer()),
                        mapping_df=writer_map,
                        raw_df=st.session_state.raw_df,
                        template_filename=st.session_state.template_file.name
                    )
                    keep_vba = st.session_state.template_file.name.lower().endswith(".xlsm")
                    final_name = sanitize_filename(output_name, keep_xlsm=keep_vba)
                    st.success("Done! Your updated masterfile is ready. Duplicate Partner SKU / Barcode cells are highlighted.")
                    st.download_button(
                        label="⬇️ Download Updated Masterfile",
                        data=out_bytes.getvalue(),
                        file_name=final_name,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
            except Exception as e:
                st.exception(e)
